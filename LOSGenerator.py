from io import StringIO
import pandas as pd
import openpyxl as opxl
from openpyxl.styles import Alignment, Border, PatternFill, Side
import re
import tkinter as tk
from tkinter import filedialog

# Colors
lt_grey = "8e8e8e"
drk_grey = "474747"
black = "ffffff"
white = "000000"

# Output file functions

# Create a row of LOS data in the specified worksheet
def _create_movement_row(ws, start_row, period, direction):

    # Create section header
    ws.cell(row=start_row, column=2, value=period).alignment = Alignment(horizontal='center')
    ws.cell(row=start_row, column=3, value=direction).alignment = Alignment(horizontal='center')

    # Copy headers from the top for data section
    for column in range(1,29):
        value = ws.cell(row=2, column=column).value
        ws.cell(row=start_row + 1, column=column, value=value).alignment = Alignment(horizontal='center')

    # Copy subheaders
    for column in range(1,29):
        value = ws.cell(row=3, column=column).value
        ws.cell(row=start_row + 2, column=column, value=value).alignment = Alignment(horizontal='center')

# Format cells with borders and typeface
def _format_cell(ws, row, col, sides='lrtb', style='thick', color="000000"):
    top, bottom, left, right = None
    if 'l' in sides:
        left = Side(border_style=style, color=color)
    if 'r' in sides:
        right = Side(border_style=style, color=color)
    if 't' in sides:
        top = Side(border_style=style, color=color)
    if 'b' in sides:
        bottom = Side(border_style=style, color=color)

    # Create and Apply border
    border = Border(top=top, left=left, right=right, bottom=bottom)
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    ws.cell(row=row, column=col).border = border
    ws.cell(row=row, column=col).fill = fill

# Creates Excel workbook and sets up initial headers
def create_workbook(initial_headers=False):
    wb = opxl.Workbook()
    if initial_headers:
        write_header_rows(wb.active, at_row=0)
    return wb

def format_borders(ws, row, col, data, sides='lrtb'):
    # Apply border
    for i in range(0,13):

        # Upper-left cell
        if row == 0 and col == 0:
            _format_cell(ws, row, col=col+i, sides='lt', style='thick')

        # Upper-right cell
        elif row == 0 and col == 15:
            _format_cell(ws, row, col=col+i, sides='tr', style='thick')

        # Top line
        elif row == 0:
            _format_cell(ws, row, col=col+i, sides='t', style='thick')

        # Bottom-left cell
        elif row == len(data) and col == 0:
            _format_cell(ws, row, col=col+i, sides='bl', style='thick')

        # Bottom-right cell
        elif row == len(data) and col == 15:
            _format_cell(ws, row, col=col+i, sides='br', style='thick')

        # Bottom line
        elif row == len(data):
            _format_cell(ws, row, col=col+i, sides='b', style='thick')

        # Left line
        elif col == 0:
            _format_cell(ws, row, col=col+i, sides='l', style='thick')

        # Right line
        elif col == 15:
            _format_cell(ws, row, col=col+i, sides='r', style='thick')

        # Inner cells
        else:
            _format_cell(ws, row, col=col+i, sides='lrtb', style='thin')

    # Advance row
    row += 1

def write_header_rows(ws, at_row):
    # Lists of Headers (Columns A-S)
    # Row 1
    headers_1 = ["Node","","Street Name","","EXISTING","","","","","","","","",
                 "Node","","Street Name","","","","","PROPOSED","","","","","","","",""]

    # Row 2
    headers_2 = ["Time","Direction","Mvmt","Link Dist","Volume","Delay","Delay","LOS","LOS","Vol%","v/c","Q50","Q95",
                 "Q95","Cycle Length", "Split","Offset","Notes", "", "Time","Direction","Mvmt","Link Dist","Volume",
                 "Delay","Delay","LOS","LOS","Vol%","v/c","Q50","Q95","Q95", "Cycle Length","Split","Offset","Notes"]

    # Row 3
    headers_3 = ["","","","","","Synchro","Synchro","SimT","Synchro","SimT","SimT","Synchro","Synchro","Synchro","SimT",
                 "","","","","Synchro","Synchro","SimT","Synchro","SimT","SimT","Synchro","Synchro","Synchro","SimT"]

    # Write line 1
    for i, header in enumerate(headers_1):
        ws.write(at_row, i, header)

    # Write line 2
    for i, header in enumerate(headers_2):
        ws.write(at_row+1, i, header)

    # Write line 3
    for i, header in enumerate(headers_3):
        ws.write(at_row+2, i, header)

# Input file and Pandas functions

def _build_signal_dataframe(name, intx_lines):

    # Signal timing dataframe
    signal_df = pd.DataFrame(name=name+' Signal', columns=['Lane Group','Lane Configurations','Cycle Length','Offset'])

    for line, i in enumerate(intx_lines):
        if line.startswith("Lane Group"):
            lane_groups = line.split('\t')[1].strip()
            for lg in lane_groups:
                signal_df[signal_df['Lane Group']] = lg
        elif line.startswith("Lane Configurations"):
            lane_configs = line.split('\t')[1].strip()
            for lc in lane_configs:
                signal_df[signal_df['Lane Configrations']] = lc
        elif line.startswith("Cycle Length"):
            cl = int(line.split(": ")[1])
            signal_df[signal_df['Cycle Length']] = cl
        elif line.startswith("Offset"):
            # Need to use regex for this one
            offset = int(re.search(r"Offset: (\d+)", line).group(1))
            signal_df[signal_df['Offset']] = offset

def _build_traffic_dataframe(name, intx_lines):
    
    # Traffic dataframe
    traffic_df = pd.DataFrame(name=name+' Traffic', columns=['Lane Group','Lane Configurations','Traffic Volume (vph)',
                                                             'Total Delay','v/c Ratio','Queue Length 50th (ft)',
                                                             'Queue Length 95th (ft)'])

    for line, i in enumerate(intx_lines):
        if line.startswith("Lane Group"):
            lane_groups = line.split('\t')[1].strip()
            for lg in lane_groups:
                traffic_df[traffic_df['Lane Group']] = lg
        elif line.startswith("Lane Configurations"):
            lane_configs = line.split('\t')[1].strip()
            for lc in lane_configs:
                traffic_df[traffic_df['Lane Configurations']] = lc
        elif line.startswith("Traffic Volume (vph)"):
            vols = line.split('\t')[1]
            for vol in vols:
                traffic_df[traffic_df['Traffic Volume (vph)']] = vol
        elif line.startswith("Total Delay"):
            delays = line.split('\t')[1]
            for delay in delays:
                traffic_df[traffic_df['Total Delay']] = delay
        elif line.startswith("v/c Ratio"):
            vc_ratios = line.split('\t')[1]
            for vc in vc_ratios:
                traffic_df[traffic_df['v/c Ratio']] = vc
        elif line.startswith("LOS"):
            los = line.split('\t')[1]
            for l in los:
                traffic_df[traffic_df['LOS']] = l
        elif line.startswith("Queue Length 50th (ft)"):
            q50s = line.split('\t')[1]
            for q in q50s:
                traffic_df[traffic_df['Queue Length 50th (ft)']] = q
        elif line.startswith("Queue Length 95th (ft)"):
            q95s = line.split('\t')[1]
            for q in q95s:
                traffic_df[traffic_df['Queue Length 95th (ft)']] = q
    
    return traffic_df

# Convenience function to do all filters at once depending on whether traffic or signal df's are desired
def _filter_empty_configs_and_lines(df, configuration='traffic'):
    int_df = _filter_empty_lane_configs(df)
    if configuration == 'traffic':
        return _filter_traffic_lines(int_df)
    elif configuration == 'signal':
        return _filter_signal_lines(int_df)
    else:
        raise ValueError("Filter type must either be <traffic> or <signal>")

def _filter_empty_lane_configs(df):
    # Filters out columns without lane config defined
    return df[df['Lane Configurations'] != '\t']

def _filter_traffic_lines(df):
    traffic_lines = ['Lane Group','Lane Configurations','Traffic Volume (vph)','v/c Ratio',
                     'Total Delay','LOS','Queue Length 50th (ft)','Queue Length 95th (ft)']
    return df[df['Lane Configurations'].isin(traffic_lines)]

def _filter_signal_lines(df):
    signal_lines = ['Cycle Length', 'Offset', 'Total Split (s)']
    return df[df['Lane Configurations'].isin(signal_lines)]

# Split lines per intersection and build a list of lines. Returns a tuple containing the node data and node name
def _split_by_node(lines):
    node_lines = []
    intersection_lines = []
    names = []
    name_line = False

    for line in lines:
        if line.startswith("Lanes, Volumes, Timings"):
            if intersection_lines:
                name_line = True # Next line is the intersection name
                node_lines.append(intersection_lines)
            if name_line:
                name_line = False
                name = line.split(":\t")[1].strip()
                names.append(name)
            intersection_lines = []
        intersection_lines.append(line)
    
    # Append last node
    if intersection_lines:
        node_lines.append(intersection_lines)

    return (node_lines, names)

# Main function
def los_generator(txt_path, xlsx_path):

    # Create workbook and select active sheet
    wb = opxl.Workbook()
    ws = wb.active

    # Create workbook with initial headers
    create_workbook(xlsx_path)

    # Read in Synchro file
    with open(txt_path, 'r') as file:
        lines = file.readlines()
    
    # Get node names and data by node
    (node_lines, names) = _split_by_node(lines)

    # Now go through node_lines for each node, and build a traffic and signal DataFrame for each
    node_df = pd.DataFrame(columns=['NodeName','TrafficDF','SignalDF'])
    for i, node in enumerate(node_lines):
        tdf = _build_traffic_dataframe(name=names[i], intx_lines=node_lines[i])
        sdf = _build_signal_dataframe(name=names[i], intx_lines=node_lines[i])
        tdf_cleaned = _filter_empty_configs_and_lines(tdf, configuration='traffic')
        sdf_cleaned = _filter_empty_configs_and_lines(sdf, configuration='signal')
        node_df[node_df['TrafficDF']] = tdf
        node_df[node_df['SignalDF']] = sdf


# Main Script

# Open file chooser dialog for Synchro report .txt
# Create and hide Tk root window
root = tk.Tk()
root.withdraw()

# Ask user for input .txt file from Synchro
input_path = filedialog.askopenfilename(title="Select Synchro report .txt file", defaultextension=".txt",
                                        filetypes=(("Text files","*.txt"),("All files","*.*")))
print(f"Synchro report file: {input_path}")

# Hide tk again
root.deiconify()

# Ask user for output location with .xlsx filename
output_path = filedialog.asksaveasfilename(title="Save output as .xlsx", defaultextension=".xlsx",
                                               filetypes=(("Excel files","*.xlsx"),("All files","*.*")))
print(f"Output excel spreadsheet: {output_path}")

root.deiconify()

# Process the input data
los_generator(input_path, output_path)