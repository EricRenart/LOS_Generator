import pytest
import openpyxl as opxl
import pandas as pd
from pandas.testing import assert_frame_equal
import LOSGenerator
import os

class LOSGeneratorTestData:
    PATH = os.path.dirname(os.path.abspath(__file__)) + "\\tests\\"
    TRAFFIC_LINES_PATH = PATH.join('test_lines_traffic.txt')
    SIGNAL_LINES_PATH = PATH.join('test_lines_signal.txt')

    # Testing data is from an actual project in Flushing I did analysis for
    NODE_NAMES = {1: "162nd Street & Northern Boulevard",
                     2: "162nd Street & Crocheron Avenue",
                     3: "Crocheron Avenue & 161st Street"}
    LANE_GROUPS = {1: ['EBL','EBT','EBR','WBL','WBT','WBR','NBL','NBT','NBR','SBL','SBT','SBR',
                       'SEL','SET','SER','NWL','NWT','NWR'],
                    2: ['EBL','EBT','EBR','WBL','WBT','WBR','NBL','NBT','NBR','SBL','SBT','SBR',
                    '�1','�5'],
                    3: ['EBL','EBT','EBR','WBL','WBT','WBR','NBL','NBT','NBR','SBL','SBT','SBR',
                    '�3','�7']}
    
    # HEaders for Excel sheet
    VALID_HEADERS = {1: ["Node", "", "Street Name", "", "EXISTING", "", "", "", "", "", "", "", "",
                    "Node", "", "Street Name", "", "", "", "", "PROPOSED", "", "", "", "", "", "", "", ""],
                     2: ["Time", "Direction", "Mvmt", "Link Dist", "Volume", "Delay", "Delay", "LOS", "LOS", "Vol%", "v/c",
                    "Q50", "Q95", "Q95", "Cycle Length", "Split", "Offset", "Notes", "", "Time", "Direction", "Mvmt", "Link Dist",
                    "Volume", "Delay", "Delay", "LOS", "LOS", "Vol%", "v/c", "Q50", "Q95", "Q95", "Cycle Length", "Split",
                    "Offset", "Notes"],
                     3: ["", "", "", "", "", "Synchro", "Synchro", "SimT", "Synchro", "SimT", "SimT", "Synchro", "Synchro",
                    "Synchro", "SimT",
                    "", "", "", "", "Synchro", "Synchro", "SimT", "Synchro", "SimT", "SimT", "Synchro", "Synchro",
                    "Synchro", "SimT"]}
    
    # Sample traffic, signal and node (intersection) dataframes
    TRAFFIC_DF_1 = pd.DataFrame(data={'Lane Groups': LANE_GROUPS[1],
                                 'Traffic Volume': [0,0,0,0,0,0,78,159,11,143,177,0,0,625,127,19,1334,26],
                                 'Total Delay': [0,0,0,0,0,0,0,127.0,0,0,31.1,0,0,13.7,0,4.2,14.2,0],
                                 'v/c Ratio': [0,0,0,0,0,0,0,0.94,0,0,0.55,0,0,0.57,0,0.08,0.70,0],
                                 'LOS': ['A','A','A','A','A','A','F','A','A','C','A','A','B','A'],
                                 'Queue 50th': ['0','0','0','0','0','0','0','219','0','0','77','0','0',
                                                '100','0','2','504'],
                                'Queue 95th': ['0','0','0','0','0','0','0','#364','0','0','102','0','0',
                                                '118','0','m4','606']},
                                name=NODE_NAMES[1]+" Traffic")

    SIGNAL_DF_1 = pd.DataFrame(data={'Cycle Length': 120,
                                      'Offset': 23,
                                      'Splits': None}, # Splits TBI
                                name=NODE_NAMES[1]+" Signal")
    
    NODE_DF_1 = pd.DataFrame(data={'Name': NODE_NAMES[1],'TrafficDF':TRAFFIC_DF_1,'SignalDF':SIGNAL_DF_1})

class LOSGeneratorTestHelpers:

    # Test Helper Functions.
    # Creates a blank workbook in the test directory with optional LOS headers. Omit .xlsx in filename.
    def create_blank_workbook(self, filename='test', initial_headers=False):
        wb = LOSGenerator.create_workbook(initial_headers=initial_headers)
        if os.path.splitext(LOSGeneratorTestData.PATH)[1] == ".xlsx": # ensure filename does not contain .xlsx
            raise ValueError("Specified filename should not contain .xlsx.")
        wb.save(os.path.join(LOSGeneratorTestData.PATH, f"{filename}.xlsx"))
        return wb

    # Deletes all files in test folder.
    def clear_test_data(self):
        for file in os.listdir(LOSGeneratorTestData.PATH):
            path = os.path.join(LOSGeneratorTestData.PATH, file)
            if os.path.isfile(path):
                os.remove(path)
    
    def test_lines_node(self):
        with open(LOSGeneratorTestData.PATH) as file:
            lines = file.readlines()
        return lines
    
    def test_lines_traffic(self):
        with open(LOSGeneratorTestData.TRAFFIC_LINES_PATH) as file:
            lines = file.readlines()
        return lines
    
    def test_lines_signal(self):
        with open(LOSGeneratorTestData.SIGNAL_LINES_PATH) as file:
            lines = file.readlines()
        return lines

class LOSGeneratorTests:
    # Test Functions

    def test_workbook_create(self):
        print('test_workbook_create()')
        fname = 'test_workbook_create'
        LOSGeneratorTestHelpers.create_blank_workbook(LOSGeneratorTestData.PATH)
        assert os.path.exists(LOSGeneratorTestData.PATH.join(f"{fname}.xlsx"))
        LOSGeneratorTestHelpers.clear_test_data()

    def test_workbook_create_with_headers(self):
        print('test_workbook_create_with_headers()')
        fname = 'test_workbook_create_with_headers'
        LOSGeneratorTestHelpers.create_blank_workbook(LOSGeneratorTestData.PATH, initial_headers=True)

        # Does workbook exist?
        assert os.path.exists(LOSGeneratorTestData.PATH.join(f"{fname}.xlsx"))

        # Do headers exist in workbook?
        wb = opxl.load_workbook(LOSGeneratorTestData.PATH.join(f"{fname}.xlsx"))
        ws = wb.active

        # Check each row in the worksheet
        for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=15):
            for cell in row:
                if cell.value not in LOSGeneratorTestData.VALID_HEADERS[0]:
                    # Fail if headers differ
                    pytest.fail()

        for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=15):
            for cell in row:
                if cell.value not in LOSGeneratorTestData.VALID_HEADERS[1]:
                    # Fail if headers differ
                    pytest.fail()

        for row in ws.iter_rows(min_row=4, max_row=4, min_col=1, max_col=15):
            for cell in row:
                if cell.value not in LOSGeneratorTestData.VALID_HEADERS[2]:
                    # Fail if headers differ
                    pytest.fail()

        LOSGeneratorTestHelpers.clear_test_data()
    
    def test_build_traffic_df(self):
        print('test_build_traffic_df()')
        traffic_lines = LOSGeneratorTestHelpers.test_lines_traffic()
        tdf = LOSGenerator._build_traffic_dataframe(name=LOSGeneratorTestData.NODE_NAMES[1], lines=traffic_lines)
        pd.testing.assert_frame_equal(tdf, LOSGeneratorTestHelpers.TRAFFIC_DF_1)
    
    def test_build_signal_df(self):
        print('test_build_signal_df()')
        signal_lines = LOSGeneratorTestHelpers.test_lines_signal()
        sdf = LOSGenerator._build_signal_dataframe(name=LOSGeneratorTestData.NODE_NAMES[1], lines=signal_lines)
        pd.testing.assert_frame_equal(sdf, LOSGeneratorTestHelpers.SIGNAL_DF_1)
    
    def test_build_node_df(self):
        print('test_build_node_df()')
        traffic_lines = LOSGeneratorTestHelpers.test_lines_traffic()
        signal_lines = LOSGeneratorTestHelpers.test_lines_signal()
        tdf = LOSGenerator._build_traffic_dataframe(name=LOSGeneratorTestData.NODE_NAMES[1], lines=traffic_lines)
        sdf = LOSGenerator._build_signal_dataframe(name=LOSGeneratorTestData.NODE_NAMES[1], lines=signal_lines)
        pd.testing.assert_frame_equal(tdf, LOSGeneratorTestHelpers.TRAFFIC_DF_1)
        pd.testing.assert_frame_equal(sdf, LOSGeneratorTestHelpers.SIGNAL_DF_1)
    
    def test_import_and_build_traffic_df(self):
        print('test_import_and_build_traffic_df()')
    
    def test_import_and_build_signal_df(self):
        print('test_import_and_build_signal_df()')
    
    def test_import_full(self):
        print('test_import_full()')
    
    def test_drop_empty_lane_groups(self):
        print('test_drop_empty_lane_groups()')
    
    def test_split_by_node(self):
        print('test_split_by_node()')